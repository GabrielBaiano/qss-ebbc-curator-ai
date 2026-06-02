import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill_sheet(ws, data, year, header_color, tab_color, zebra_color):
    # Definir propriedades da aba
    ws.title = year
    ws.sheet_properties.tabColor = tab_color
    
    # Cabeçalhos originais da planilha
    headers = [
        "DOI",
        "Título",
        "Autoria",
        "Palavras-chave",
        "Ferramenta utilizada",
        "Identifica a ferramenta?",
        "Onde usou (coleta dos dados, análise dos dados ou visualização - gerar gráficos)",
        "Fonte de coleta de dados (da onde o pesquisador tirou a informação?)"
    ]
    
    # Escrever cabeçalhos
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="medium", color="1E293B"),
        bottom=Side(style="medium", color="1E293B")
    )
    
    # Aplicar estilo de cabeçalho na linha 1
    ws.row_dimensions[1].height = 40
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        
    # Estilos de dados
    font_data = Font(name="Segoe UI", size=10, color="333333")
    font_link = Font(name="Segoe UI", size=10, underline="single", color="2563EB")
    
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
    
    border_cell = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Preencher dados
    for row_idx, item in enumerate(data, start=2):
        # Mapeamento do item JSON para as colunas
        row_data = [
            item.get("DOI", ""),
            item.get("Título", ""),
            item.get("Autoria", ""),
            item.get("Palavras-chave", ""),
            item.get("Ferramenta utilizada", ""),
            item.get("Identifica a ferramenta?", ""),
            item.get("Onde usou (coleta dos dados, análise dos dados ou visualização - gerar gráficos)", ""),
            item.get("Fonte de coleta de dados (da onde o pesquisador tirou a informação?)", "")
        ]
        
        # Inserir linha
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 24  # Altura confortável para leitura com wrap_text
        
        # Cor de fundo alternada (zebra striping)
        current_fill = fill_zebra if row_idx % 2 == 0 else fill_white
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            cell.border = border_cell
            
            # Centralizar coluna 'Identifica a ferramenta?' e alinhar as outras à esquerda
            if col_idx == 6:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Tratar coluna DOI (coluna 1) como hiperlink real
            if col_idx == 1 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = font_link
            else:
                cell.font = font_data
                
    # Congelar cabeçalho (linha 1)
    ws.freeze_panes = "A2"
    
    # Adicionar filtros automáticos nas colunas do cabeçalho
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
    
    # Configurar larguras ideais de colunas
    col_widths = {
        1: 28,  # DOI
        2: 45,  # Título
        3: 32,  # Autoria
        4: 32,  # Palavras-chave
        5: 25,  # Ferramenta utilizada
        6: 15,  # Identifica a ferramenta?
        7: 25,  # Onde usou
        8: 30   # Fonte de coleta de dados
    }
    
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

def main():
    print("Iniciando a geração e estilização da planilha excel...")
    
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Carregar dados do Volume 5 (2024)
    vol5_path = os.path.join(root_dir, "datasets", "qss_volume_5_data.json")
    if not os.path.exists(vol5_path):
        raise FileNotFoundError(f"Arquivo {vol5_path} não encontrado!")
    with open(vol5_path, "r", encoding="utf-8") as f:
        data_2024 = json.load(f)
        
    # Carregar dados do Volume 6 (2025)
    vol6_path = os.path.join(root_dir, "datasets", "qss_volume_6_data.json")
    if not os.path.exists(vol6_path):
        raise FileNotFoundError(f"Arquivo {vol6_path} não encontrado!")
    with open(vol6_path, "r", encoding="utf-8") as f:
        data_2025 = json.load(f)
        
    print(f"Dados carregados: {len(data_2024)} artigos para 2024 e {len(data_2025)} artigos para 2025.")
    
    # Criar novo workbook
    wb = Workbook()
    
    # Configurar aba 2024 (Planilha 1)
    ws_2024 = wb.active
    # Cabeçalho Slate `#1E293B`, Tab Slate `#1E293B`, Zebra suave `#F8FAFC`
    fill_sheet(ws_2024, data_2024, "2024", "1E293B", "1E293B", "F8FAFC")
    
    # Configurar aba 2025 (Planilha 2)
    ws_2025 = wb.create_sheet()
    # Cabeçalho Deep Teal `#0F766E`, Tab Deep Teal `#0F766E`, Zebra suave `#F0FDF4`
    fill_sheet(ws_2025, data_2025, "2025", "0F766E", "0F766E", "F0FDF4")
    
    # Salvar a planilha no caminho especificado
    excel_output_path = os.path.join(root_dir, "coleta de dados gabriel.xlsx")
    wb.save(excel_output_path)
    
    print(f"Sucesso! A planilha '{excel_output_path}' foi criada e estilizada com maestria.")

if __name__ == "__main__":
    main()
