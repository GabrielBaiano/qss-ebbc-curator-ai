import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fill_sheet(ws, data, title, header_color, tab_color, zebra_color):
    # Definir propriedades da aba
    ws.title = title
    ws.sheet_properties.tabColor = tab_color
    
    # Cabeçalhos originais da planilha
    headers = [
        "DOI",
        "Título",
        "Autoria",
        "Palavras-chave",
        "Ferramenta utilizada",
        "Identifica a ferramenta?",
        "Onde usou (coleta dos dados, análise dos dados ou visualização - gerar gráficos)",
        "Fonte de coleta de dados (da onde o pesquisador tirou a informação?)"
    ]
    
    # Escrever cabeçalhos
    ws.append(headers)
    
    # Estilizar cabeçalho
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="medium", color="1E293B"),
        bottom=Side(style="medium", color="1E293B")
    )
    
    # Aplicar estilo de cabeçalho na linha 1
    ws.row_dimensions[1].height = 40
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        
    # Estilos de dados
    font_data = Font(name="Segoe UI", size=10, color="333333")
    font_link = Font(name="Segoe UI", size=10, underline="single", color="2563EB")
    
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
    
    border_cell = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Preencher dados
    for row_idx, item in enumerate(data, start=2):
        row_data = [
            item.get("DOI", ""),
            item.get("Título", ""),
            item.get("Autoria", ""),
            item.get("Palavras-chave", ""),
            item.get("Ferramenta utilizada", ""),
            item.get("Identifica a ferramenta?", ""),
            item.get("Onde usou (coleta dos dados, análise dos dados ou visualização - gerar gráficos)", ""),
            item.get("Fonte de coleta de dados (da onde o pesquisador tirou a informação?)", "")
        ]
        
        # Inserir linha
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 24
        
        # Cor de fundo alternada (zebra striping)
        current_fill = fill_zebra if row_idx % 2 == 0 else fill_white
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            cell.border = border_cell
            
            # Centralizar coluna 'Identifica a ferramenta?' e alinhar as outras à esquerda
            if col_idx == 6:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Tratar coluna DOI (coluna 1) como hiperlink real se for URL válida
            if col_idx == 1 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = font_link
            else:
                cell.font = font_data
                
    # Congelar cabeçalho (linha 1)
    ws.freeze_panes = "A2"
    
    # Adicionar filtros automáticos
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
    
    # Configurar larguras ideais de colunas
    col_widths = {
        1: 28,  # DOI
        2: 45,  # Título
        3: 32,  # Autoria
        4: 32,  # Palavras-chave
        5: 25,  # Ferramenta utilizada
        6: 15,  # Identifica a ferramenta?
        7: 25,  # Onde usou
        8: 30   # Fonte de coleta de dados
    }
    
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

def load_json(filename):
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    filepath = os.path.join(root_dir, "datasets", filename)
    if not os.path.exists(filepath):
        print(f"Aviso: Arquivo {filepath} não encontrado! Aba correspondente será criada vazia.")
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)

def main():
    print("Iniciando a consolidação estilizada da planilha Excel...")
    
    wb = Workbook()
    
    # Configuração dos dados de entrada: (nome_arquivo, nome_da_aba, cor_cabecalho, cor_tab, cor_zebra)
    sheets_config = [
        # QSS
        ("qss_volume_1_data.json", "QSS 2020", "1E3A8A", "1E3A8A", "F0F9FF"),
        ("qss_volume_2_data.json", "QSS 2021", "B45309", "B45309", "FEF3C7"),
        ("qss_volume_3_data.json", "QSS 2022", "6D28D9", "6D28D9", "F5F3FF"),
        ("qss_volume_4_data.json", "QSS 2023", "BE185D", "BE185D", "FDF2F8"),
        ("qss_volume_5_data.json", "QSS 2024", "1E293B", "1E293B", "F8FAFC"),
        ("qss_volume_6_data.json", "QSS 2025", "0F766E", "0F766E", "F0FDF4"),
        # EBBC
        ("ebbc_2020_data.json", "EBBC 2020", "047857", "047857", "ECFDF5"),
        ("ebbc_2022_data.json", "EBBC 2022", "C2410C", "C2410C", "FFF7ED"),
        ("ebbc_2024_data.json", "EBBC 2024", "4338CA", "4338CA", "EEF2FF")
    ]
    
    first = True
    for filename, tab_title, header_col, tab_col, zebra_col in sheets_config:
        print(f"Carregando dados para aba: {tab_title}...")
        data = load_json(filename)
        print(f"Aba '{tab_title}' carregada com {len(data)} registros.")
        
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
            
        fill_sheet(ws, data, tab_title, header_col, tab_col, zebra_col)
        
    root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_output_path = os.path.join(root_dir, "coleta de dados gabriel.xlsx")
    wb.save(excel_output_path)
    print(f"\nSucesso! A planilha '{excel_output_path}' foi atualizada e estilizada com maestria.")

if __name__ == "__main__":
    main()
